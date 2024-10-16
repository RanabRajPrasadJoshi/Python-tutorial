from fileinput import filename

import openpyxl as xl
import openpyxl.utils.exceptions

from pathlib import Path


def primenumber_finder(number1):
    count = 0
    for num in range(1, number1 + 1):
        if(number1%num == 0):
            count +=1

    if(count>2):
        return "not prime"
    else:
        return "prime"

def excel_updater(filename , sheetname ):
    wb = xl.load_workbook(filename)
    sheet = wb[sheetname]
    for value in range(2, sheet.max_row + 1):
        number = sheet.cell(value , 1).value
        result=primenumber_finder(number)
        result_cell = sheet.cell(value , 2)
        result_cell.value = result
    flag = input("\nTo save as per requirement 'Y' for create new one and 'N' for override it (Y/N) \n")
    try:
        if flag.upper() == "Y":
            new_file = input("Enter the new file name:")
            wb.save(new_file)
            print("Successfully new file created and updated  please check it  by openaing the file")
        elif flag.upper == "N":
            wb.save(filename)
            print("Successfully updated please check it openaing the file you updated")

    except(PermissionError):
        print("your file nust be opened please close if else something wrong accored try again")


try:
    filename = input("Enter the file name with extension exactly as it is: ")
    path = Path(filename)
    if path.exists():
        sheet = input("Enter the sheet name exactly as it is: ")

        excel_updater(filename , sheet)
    else:
        print("File doesnot exists in the directory")


except(openpyxl.utils.exceptions.InvalidFileException):
    print("Invalid file extension try again ('_') ")
except(FileNotFoundError):
    print("Make sure that the file and this program must be in same folder and the file name and extension should be correct ('_')")
except(KeyError):
    print("Sheet name must be incorrect ('_') ")








