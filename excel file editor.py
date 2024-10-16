from dis import CACHE

import openpyxl as xl

wb = xl.load_workbook('transaction.xlsx')
sheet = wb ['Sheet2']

for value in range(2,sheet.max_row + 1):
    cell = sheet.cell(value , 1)
    corrected_value = cell.value + 20
    corrected_cell = sheet.cell(value , 4)
    corrected_cell.value = corrected_value
try:
    wb.save('transaction.xlsx')
except(PermissionError):
    print("Close the file which should be edited id not solved some error occured try again or contach to customer team")
